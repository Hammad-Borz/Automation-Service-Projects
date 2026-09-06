"""Formatted Excel workbook export."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from .models import AnalyticsResult


class ExcelExporter:
    """Create a multi-sheet workbook with summary tables and charts."""

    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")

    def export(self, analytics: AnalyticsResult, destination: Path) -> Path:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Executive Summary"
        summary.append(["ReportFlow Business Report"])
        summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        summary["A1"].fill = self.HEADER_FILL
        summary.merge_cells("A1:B1")
        summary.append([])
        for key, value in analytics.kpis.items():
            summary.append([key.replace("_", " ").title(), value])
        self._write_table(workbook, "Products", analytics.by_product)
        self._write_table(workbook, "Categories", analytics.by_category)
        self._write_table(workbook, "Regions", analytics.by_region)
        self._write_table(workbook, "Monthly Trends", analytics.monthly)
        self._add_charts(workbook["Products"], workbook["Monthly Trends"])
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.column_dimensions["A"].width = 25
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return destination

    @staticmethod
    def _write_table(workbook: Workbook, name: str, frame) -> None:
        sheet = workbook.create_sheet(name)
        sheet.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            sheet.append(list(row))
        for column in ("revenue", "unit_price"):
            if column in frame.columns:
                index = list(frame.columns).index(column) + 1
                for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
                    for value in cell:
                        value.number_format = '$#,##0.00'

    @staticmethod
    def _add_charts(products, monthly) -> None:
        product_chart = BarChart()
        product_chart.title = "Revenue by Product"
        product_chart.add_data(Reference(products, min_col=2, min_row=1, max_row=products.max_row), titles_from_data=True)
        product_chart.set_categories(Reference(products, min_col=1, min_row=2, max_row=products.max_row))
        products.add_chart(product_chart, "E2")
        trend_chart = LineChart()
        trend_chart.title = "Monthly Revenue Trend"
        trend_chart.add_data(Reference(monthly, min_col=2, min_row=1, max_row=monthly.max_row), titles_from_data=True)
        trend_chart.set_categories(Reference(monthly, min_col=1, min_row=2, max_row=monthly.max_row))
        monthly.add_chart(trend_chart, "E2")
