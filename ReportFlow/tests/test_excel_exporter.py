from openpyxl import load_workbook

from src.analytics import Analytics
from src.data_processor import DataProcessor
from src.excel_exporter import ExcelExporter


def test_excel_export_has_sheets_and_charts(sample_frame, tmp_path) -> None:
    analytics = Analytics().calculate(DataProcessor().process(sample_frame))
    path = ExcelExporter().export(analytics, tmp_path / "report.xlsx")
    workbook = load_workbook(path)
    assert {"Executive Summary", "Products", "Regions", "Monthly Trends"}.issubset(workbook.sheetnames)
    assert len(workbook["Products"]._charts) == 1
